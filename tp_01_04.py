import openpyxl as excel
import tkinter as tk
import tkinter.messagebox as msg
import tkinter.filedialog as fl
import traceback
import sys
import math

def readExcel():
    '''READ_Excel'''
    excelWs=[
        excel.load_workbook(fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'tp*_01.xlsx')],title="select [tp*_01.xlsx ]file")).worksheets[0],
        excel.load_workbook(fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'tp*_02.xlsx')],title="select [tp*_02.xlsx ]file")).worksheets[0],
        excel.load_workbook(fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'tp*_03.xlsx')],title="select [tp*_03.xlsx ]file")).worksheets[0],
        excel.load_workbook(fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'tp*_04.xlsx')],title="select [tp*_04.xlsx ]file")).worksheets[0]
    ]

    cdChange={
        "内用薬":"1",
        "注射薬":"2",
        "外用薬":"3",
        "歯科用薬剤":"4"
    }

    arr=[]
    qeuryArr=[]

    for ws in excelWs:
        # arr.append("INSERT INTO `medicine_generic_mst` VALUES ")
        for row in ws.iter_rows(min_col=1,min_row=2, max_col=9, max_row=ws.max_row, values_only=True):
            add_comma="','".join([cdChange.get(row[0]),row[1],row[2],row[3],row[7],'' if row[8] is None else str(row[8])])
            arr.append(",('"+add_comma+"')")
    

    #CREATE LIST
    while True: 
        num=len(qeuryArr)+1
        rowcnt_devided=math.ceil(len(arr)/5000)-1

        add_5000=arr[(num-1)*5000:num*5000]
        add_5000_comma="".join(add_5000).replace(',','',1)
        qeuryArr.append("INSERT INTO `medicine_price_mst` VALUES "+add_5000_comma+";\n")
        
        if rowcnt_devided<num:              
            break
        
    return qeuryArr

def writeExcel(qeuryArr):
    '''FIRST+MAIN+LAST SQL'''
    first_stream=""
    main_stream=''.join(qeuryArr)
    last_stream=""

    #FIRST FILE
    fileName_first=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_price_mst_template_first.sql')],title="select [sos_medicine_price_mst_template_first.sql] file")
    with open(fileName_first, 'r', encoding="utf-8") as firstFile:
        first_stream = firstFile.read()

    #LAST FILE
    fileName_last=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_price_mst_template_last.sql')],title="select [sos_medicine_price_mst_template_last.sql] file")
    with open(fileName_last, 'r', encoding="utf-8") as lastFile:
        last_stream = lastFile.read()

    #COMBINE FILE
    with open('./sos_medicine_price_mst.sql', 'w',-1,'utf-8') as f:
        f.write(first_stream+"\n")
        f.write(main_stream+"\n")
        f.write(last_stream+"\n")

def start():
    try:
        writeExcel(readExcel())
        msg.showinfo(title='成功', message='保存しました。')
    except (excel.utils.exceptions.InvalidFileException , FileNotFoundError):
        pass
    except Exception: 
        print(traceback.print_exc())
        msg.showerror(title='エラー', message='エラー履歴を確認してください。')