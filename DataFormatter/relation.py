import openpyxl as excel
import csv
import tkinter as tk
import tkinter.messagebox as msg
import tkinter.filedialog as fl
import traceback
import sys

def readExcel():
    '''READ_Excel'''
    fileName=fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'ippanmeishohoumaster*.xlsx')],title="select [ippanmeishohoumaster_*.xlsx ]file")
    rb = excel.load_workbook(fileName)
    ws1 = rb.worksheets[0]#一般名処方マスタ（R2.12.11版） 全体
    ws2 = rb.worksheets[1]#例外コード品目対照表
    
    arr_ippanCd=[]
    arr=[]
    befRow=""

    for row in ws1.iter_rows(min_col=2,min_row=4, max_col=2, max_row=ws1.max_row, values_only=True):
        arr_ippanCd.append(row[0])

    for row in ws2.iter_rows(min_col=2,min_row=4, max_col=6, max_row=ws2.max_row, values_only=True):
        #空白セルの場合上のセルの内容を入れる
        if row[0] is not None:
            befRow=row[0]
        arr.append([row[4],befRow])

    '''READ_CSV'''
    fileName=fl.askopenfilename(initialdir="./resource",filetypes=[('csv file', '.csv')],title="select [y_ALL*.csv ]file")
    with open(fileName,'r') as f:
        reader = csv.reader(f) 
        #ファイルのA列の最終行の次の行から貼り付け
        for row in reader:
            #A列の途中で空白あればその行を削除する
            if row[31] != '':
                arr.append([row[31],row[31][:-3]+'ZZZ'])
    
    #B列⑦で変換したデーターの中、シート「一般名処方マスタ～　全体」に存在しないデーターは全部行削除
    arr_comp=[]
    # for x in arr:
    #     if x[1] in arr_ippanCd:
    #         arr_comp.append([x[0],x[1]])
    [arr_comp.append([x[0],x[1]]) for x in arr if x[1] in arr_ippanCd]

    #重複削除機能でA列重複した行を削除する
    arr_dupl=["INSERT INTO `medicine_cd_relation` VALUES "]
    arr_dupl_check=[]
    for x in arr_comp:
        if x[0] not in arr_dupl_check:
            arr_dupl_check.append(x[0])
            arr_dupl.append(",('"+x[0]+"','"+x[1]+"')")
    arr_dupl.append(";\n")
    arr_dupl="".join(arr_dupl).replace(',','',1)  

    return arr_dupl


def writeExcel(qeuryArr):
    '''FIRST+MAIN+LAST SQL'''
    first_stream=""
    last_stream=""
    main_stream=qeuryArr

    #FIRST FILE
    fileName_first=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_cd_relation_template_first.sql')],title="select [sos_medicine_cd_relation_template_first.sql] file")
    with open(fileName_first, 'r', encoding="utf-8") as firstFile:
        first_stream = firstFile.read()

    #LAST FILE
    fileName_last=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_cd_relation_template_last.sql')],title="select [sos_medicine_cd_relation_template_last.sql] file")
    with open(fileName_last, 'r', encoding="utf-8") as lastFile:
        last_stream = lastFile.read()

    #COMBINE FILE
    with open('./sos_medicine_cd_relation.sql', 'w',-1,'utf-8') as f:
        f.write(first_stream+"")
        f.write(main_stream+"\n")
        f.write(last_stream+"\n")

def start():
    try:
        # readExcel()
        writeExcel(readExcel())
        msg.showinfo(title='成功', message='保存しました。')
    except (excel.utils.exceptions.InvalidFileException , FileNotFoundError):
        pass
    except Exception: 
        print(traceback.print_exc())
        msg.showerror(title='エラー', message='エラー履歴を確認してください。')