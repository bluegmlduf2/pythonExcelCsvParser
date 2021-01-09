import openpyxl as excel
import tkinter as tk
import tkinter.filedialog as fl
import tkinter.messagebox as msg
import traceback
import sys

def readExcel():
    '''READ_Excel'''
    fileName=fl.askopenfilename(initialdir="./resource",filetypes=[('excel file', 'ippanmeishohoumaster*.xlsx')],title="select [ippanmeishohoumaster_*.xlsx ]file")
    rb = excel.load_workbook(fileName)
    ws = rb.worksheets[0]

    arr=[]
    arr.append("INSERT INTO `medicine_common_nm_mst` VALUES ")
    for row in ws.iter_rows(min_col=2,min_row=4, max_col=5, max_row=ws.max_row, values_only=True):
        add_comma="','".join([row[0],row[2],row[3],row[1]])
        add_comma=add_comma.replace('\n','　')#\n削除
        # add_comma_2=",".join(add_comma)
        arr.append(",('"+add_comma+"')")
    arr.append(";\n")
    lastArr="".join(arr).replace(',','',1)

    return lastArr

def writeExcel(qeuryArr):
    '''FIRST+MAIN+LAST SQL'''
    first_stream=""
    last_stream=""
    main_stream=qeuryArr

    #FIRST FILE
    fileName_first=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_common_nm_mst_template_first.sql')],title="select [sos_medicine_common_nm_mst_template_first.sql] file")
    with open(fileName_first, 'r', encoding="utf-8") as firstFile:
        first_stream = firstFile.read()

    #LAST FILE
    fileName_last=fl.askopenfilename(initialdir="./templates",filetypes=[('sql file', 'sos_medicine_common_nm_mst_template_last.sql')],title="select [sos_medicine_common_nm_mst_template_last.sql] file")
    with open(fileName_last, 'r', encoding="utf-8") as lastFile:
        last_stream = lastFile.read()

    #COMBINE FILE
    with open('./sos_medicine_common_nm_mst.sql', 'w',-1,'utf-8') as f:
        f.write(first_stream+"\n")
        f.write(main_stream+"\n")
        f.write(last_stream+"\n")

def start():
    try:
        writeExcel(readExcel())
        msg.showinfo(title='成功', message='保存しました。')
    except (excel.utils.exceptions.InvalidFileException , FileNotFoundError):
        pass 
    except Exception:
        print(traceback.print_exc())
        msg.showerror(title='エラー', message='エラー履歴を確認してください。')
